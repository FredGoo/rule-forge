"""
Excel处理工具
支持读取Excel模板和写入生成的数据
"""
import os
from typing import Dict, List, Optional, Any, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelHandler:
    """Excel处理器"""
    
    def __init__(self, file_path: str):
        """初始化Excel处理器
        
        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.sheets_info = {}
    
    def read_template(self) -> Dict[str, List[str]]:
        """读取Excel模板，获取每个sheet的字段信息
        
        Returns:
            字典，key为sheet名称（类名），value为字段名列表
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")
        
        try:
            # 使用pandas读取所有sheet
            excel_file = pd.ExcelFile(self.file_path)
            sheets_info = {}
            
            for sheet_name in excel_file.sheet_names:
                # 读取第一行作为字段名
                df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)
                field_names = df.columns.tolist()
                sheets_info[sheet_name] = field_names
                
            self.sheets_info = sheets_info
            return sheets_info
            
        except Exception as e:
            raise Exception(f"读取Excel模板失败: {str(e)}")
    
    def write_data(self, data_dict: Dict[str, List[Dict[str, Any]]], 
                   output_path: Optional[str] = None) -> str:
        """将数据写入Excel文件
        
        Args:
            data_dict: 数据字典，key为sheet名称，value为数据列表
            output_path: 输出文件路径，如果为None则覆盖原文件
            
        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = self.file_path
        
        try:
            # 创建输出目录（如有需要）
            out_dir = os.path.dirname(output_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            
            # 创建新的工作簿
            workbook = Workbook()
            # 删除默认的sheet
            workbook.remove(workbook.active)
            
            for sheet_name, data_list in data_dict.items():
                if not data_list:
                    continue
                
                # 创建新的worksheet
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # 转换为DataFrame
                df = pd.DataFrame(data_list)
                
                # 写入数据
                for row in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(row)
                
                # 自动调整列宽
                self._auto_adjust_column_width(worksheet)
            
            # 保存文件
            workbook.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"写入Excel文件失败: {str(e)}")
    
    def append_data(self, sheet_name: str, data_list: List[Dict[str, Any]]) -> None:
        """向指定sheet追加数据
        
        Args:
            sheet_name: sheet名称
            data_list: 要追加的数据列表
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")
        
        try:
            # 加载现有工作簿
            workbook = load_workbook(self.file_path)
            
            if sheet_name not in workbook.sheetnames:
                # 如果sheet不存在，创建新的
                worksheet = workbook.create_sheet(title=sheet_name)
                # 写入表头
                if data_list:
                    headers = list(data_list[0].keys())
                    worksheet.append(headers)
            else:
                worksheet = workbook[sheet_name]
            
            # 追加数据
            for data_row in data_list:
                row_values = list(data_row.values())
                worksheet.append(row_values)
            
            # 自动调整列宽
            self._auto_adjust_column_width(worksheet)
            
            # 保存文件
            workbook.save(self.file_path)
            
        except Exception as e:
            raise Exception(f"追加数据到Excel失败: {str(e)}")
    
    def create_template(self, sheets_config: Dict[str, List[str]], 
                       output_path: Optional[str] = None) -> str:
        """创建Excel模板
        
        Args:
            sheets_config: sheet配置，key为sheet名称，value为字段名列表
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = self.file_path
        
        try:
            # 创建输出目录（如有需要）
            out_dir = os.path.dirname(output_path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)
            
            workbook = Workbook()
            # 删除默认的sheet
            workbook.remove(workbook.active)
            
            for sheet_name, field_names in sheets_config.items():
                # 创建新的worksheet
                worksheet = workbook.create_sheet(title=sheet_name)
                
                # 写入表头
                if field_names:
                    worksheet.append(field_names)
                
                # 自动调整列宽
                self._auto_adjust_column_width(worksheet)
            
            # 保存文件
            workbook.save(output_path)
            return output_path
            
        except Exception as e:
            raise Exception(f"创建Excel模板失败: {str(e)}")
    
    def get_sheet_names(self) -> List[str]:
        """获取所有sheet名称"""
        if not self.sheets_info:
            self.read_template()
        return list(self.sheets_info.keys())
    
    def get_field_names(self, sheet_name: str) -> List[str]:
        """获取指定sheet的字段名"""
        if not self.sheets_info:
            self.read_template()
        return self.sheets_info.get(sheet_name, [])
    
    def _auto_adjust_column_width(self, worksheet) -> None:
        """自动调整列宽"""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # 设置列宽，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception:
            # 如果调整列宽失败，忽略错误
            pass
    
    def validate_template(self) -> Tuple[bool, List[str]]:
        """验证模板格式
        
        Returns:
            (是否有效, 错误信息列表)
        """
        errors = []
        
        try:
            if not os.path.exists(self.file_path):
                errors.append(f"Excel文件不存在: {self.file_path}")
                return False, errors
            
            sheets_info = self.read_template()
            
            if not sheets_info:
                errors.append("Excel文件中没有找到任何sheet")
                return False, errors
            
            for sheet_name, field_names in sheets_info.items():
                if not field_names:
                    errors.append(f"Sheet '{sheet_name}' 没有字段定义")
                
                # 检查字段名是否有重复
                if len(field_names) != len(set(field_names)):
                    duplicates = [name for name in field_names if field_names.count(name) > 1]
                    errors.append(f"Sheet '{sheet_name}' 存在重复字段: {duplicates}")
            
            return len(errors) == 0, errors
            
        except Exception as e:
            errors.append(f"验证模板时发生错误: {str(e)}")
            return False, errors


class ExcelTemplateGenerator:
    """Excel模板生成器"""
    
    @staticmethod
    def generate_sample_template(output_path: str) -> str:
        """生成示例模板
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
        """
        sample_config = {
            "User": ["id", "username", "email", "phone", "age", "gender", "status", "createTime"],
            "Product": ["id", "name", "price", "category", "stock", "description", "isActive"],
            "Order": ["id", "orderNo", "userId", "totalAmount", "status", "createTime"],
            "Address": ["id", "province", "city", "district", "street", "zipCode", "isDefault"]
        }
        
        handler = ExcelHandler(output_path)
        return handler.create_template(sample_config, output_path)
    
    @staticmethod
    def generate_from_config(template_data: Dict[str, List[str]], output_path: str) -> str:
        """根据配置生成模板文件"""
        handler = ExcelHandler(output_path)
        return handler.create_template(template_data, output_path)