#!/usr/bin/env python3
"""
创建示例Excel模板文件
"""
import os
from openpyxl import Workbook

def create_sample_template():
    """创建示例Excel模板"""
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的工作表
    wb.remove(wb.active)
    
    # 创建User工作表
    user_sheet = wb.create_sheet("User")
    user_headers = ["id", "name", "email", "phone", "age", "status", "createTime"]
    for col, header in enumerate(user_headers, 1):
        user_sheet.cell(row=1, column=col, value=header)
    
    # 创建Product工作表
    product_sheet = wb.create_sheet("Product")
    product_headers = ["id", "name", "price", "category", "stock", "description", "isActive"]
    for col, header in enumerate(product_headers, 1):
        product_sheet.cell(row=1, column=col, value=header)
    
    # 创建Order工作表
    order_sheet = wb.create_sheet("Order")
    order_headers = ["id", "userId", "productId", "quantity", "totalAmount", "status", "orderTime"]
    for col, header in enumerate(order_headers, 1):
        order_sheet.cell(row=1, column=col, value=header)
    
    # 创建Address工作表
    address_sheet = wb.create_sheet("Address")
    address_headers = ["id", "userId", "province", "city", "district", "street", "zipCode", "isDefault"]
    for col, header in enumerate(address_headers, 1):
        address_sheet.cell(row=1, column=col, value=header)
    
    # 确保templates目录存在
    templates_dir = "templates"
    os.makedirs(templates_dir, exist_ok=True)
    
    # 保存文件
    output_path = os.path.join(templates_dir, "sample_template.xlsx")
    wb.save(output_path)
    print(f"示例模板已创建: {output_path}")

if __name__ == "__main__":
    create_sample_template()